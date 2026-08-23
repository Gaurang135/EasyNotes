from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook
from app.models import ParsedDoc, TextBlock, Table
from app.errors import CorruptFileError, NoExtractableTextError


class XlsxParser:
    file_types = frozenset({"xlsx"})

    def parse(self, path: Path) -> ParsedDoc:
        try:
            wb = load_workbook(str(path), read_only=True, data_only=True)
        except Exception as e:
            raise CorruptFileError(f"unreadable XLSX: {e}")
        blocks: list[TextBlock] = []
        tables: list[Table] = []
        try:
            for ws in wb.worksheets:
                grid = []
                for row in ws.iter_rows(values_only=True):
                    cells = ["" if v is None else str(v) for v in row]
                    if any(cells):
                        grid.append(cells)
                if not grid:
                    continue
                loc = f"{ws.title} rows 1-{len(grid)}"
                blocks.append(TextBlock(text="\n".join(",".join(r) for r in grid),
                                        kind="table", location=loc))
                tables.append(Table(name=ws.title, columns=grid[0], rows=grid[1:], location=loc))
        finally:
            wb.close()
        if not blocks:
            raise NoExtractableTextError("spreadsheet has no data")
        return ParsedDoc(text_blocks=blocks, metadata={}, warnings=[], tables=tables)
